import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ExcelExporter:
    """Excel导出器"""

    def __init__(self, output_path):
        self.output_path = output_path

    def export_results(self, results_df, _, __, ___):
        """导出仿真结果到Excel"""
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            self._create_simple_sheet(writer, results_df)

        self._format_workbook(self.output_path)
        print(f"导出完成: {self.output_path}")

    def _create_simple_sheet(self, writer, results_df):
        """创建简化数据工作表，仅包含时间和发电量"""
        simple_columns = {
            'datetime': '时间',
            'energy_kwh': '发电量(kWh)',
        }

        simple_df = results_df[list(simple_columns.keys())].copy()

        if 'energy_kwh' in simple_df.columns:
            simple_df['energy_kwh'] = simple_df['energy_kwh'].fillna(0).astype(float)

        if 'datetime' in simple_df.columns:
            simple_df['datetime'] = simple_df['datetime'].dt.tz_localize(None)

        simple_df.columns = list(simple_columns.values())
        simple_df.to_excel(writer, sheet_name='发电量预测', index=False)

    def _format_workbook(self, file_path):
        """格式Excel"""
        wb = openpyxl.load_workbook(file_path)

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        cell_font = Font(name='微软雅黑', size=10)
        cell_alignment = Alignment(horizontal='right', vertical='center')

        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.font = cell_font
                    if cell.column and cell.column > 1:
                        cell.alignment = cell_alignment
                    cell.border = border

            for column in ws.columns:
                max_length = 0
                first_cell = column[0]
                if hasattr(first_cell, 'column_letter'):
                    column_letter = first_cell.column_letter
                else:
                    continue

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            ws.freeze_panes = 'A2'

        wb.save(file_path)


def export_simulation_results(results_df, simulator, output_path):
    """导出仿真结果"""
    exporter = ExcelExporter(output_path)
    exporter.export_results(results_df, None, None, None)
